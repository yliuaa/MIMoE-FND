import random

import cv2

import torch
import torch.utils.data as data
import data.util as util

import torchvision.transforms.functional as F

from PIL import Image
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from tqdm import tqdm
from transformers import BertModel, BertTokenizer
import clip
import torchvision
import torchvision.transforms as transforms
import torch.nn as nn


class FakeNet_dataset(data.Dataset):

    def __init__(
        self,
        root_path="/home/yifan40/multimodalFNDData/data/AAAI_dataset",
        dataset="gossip",
        image_size=224,
        is_train=True,
        data_augment=False,
        duplicate_fake_times=0,
    ):
        self.duplicate_fake_times = duplicate_fake_times
        self.data_augment = data_augment
        self.dataset_name = dataset
        assert (
            self.dataset_name == "politi" or self.dataset_name == "gossip"
        ), "Error! Only 'gossip' or 'politi' supported!"
        super(FakeNet_dataset, self).__init__()
        print("duplicate_fake_times: {}".format(self.duplicate_fake_times))
        print("Dataset: {}".format(self.dataset_name))
        print("Using More Negative Examples: {}".format(self.data_augment))
        print("We are resampling bad examples using randint")
        self.is_train = is_train
        self.root_path = root_path
        self.index = 0
        self.text_max_len = 170
        # self.token = BertTokenizer.from_pretrained('bert-base-uncased')
        self.label_dict, self.label_ambiguity = [], []
        self.image_size = image_size
        self.resize_and_to_tensor = transforms.Compose(
            [
                transforms.Resize((224, 224)),
                transforms.ToTensor(),
            ]
        )
        # dataset_names = ['gossip','politi']
        # for dataset_name in dataset_names:
        # wb = openpyxl.load_workbook(root_path+'/{}_{}.xlsx'.format(dataset_name, 'train' if is_train else 'test'))
        workbook_name = self.root_path + "/{}_{}.xlsx".format(
            self.dataset_name, "train" if self.is_train else "test"
        )
        wb = openpyxl.load_workbook(workbook_name)
        print(f"Workbook name {workbook_name}")
        sheetnames = wb.sheetnames
        sheet = wb[sheetnames[0]]
        rows = sheet.max_row

        fake_news_num = 0
        ## calculate true:fake
        for i in tqdm(range(2, rows + 1)):
            images_name = str(sheet["C" + str(i)].value)
            label = int(sheet["D" + str(i)].value)
            label = 1 if label == 0 else 0
            fake_news_num += label

        thresh = (rows - fake_news_num) / (rows)
        print(f"real news: {rows-fake_news_num}")
        print(f"fake news: {fake_news_num}")
        print(f"thresh: {thresh}")
        self.pos_weight = torch.tensor((rows - fake_news_num) / fake_news_num)
        self.thresh = thresh

        skipped_num = 0
        for i in tqdm(range(2, rows + 1)):
            images_name = str(sheet["C" + str(i)].value)
            label = int(sheet["D" + str(i)].value)
            label = 1 if label == 0 else 0
            content = str(sheet["B" + str(i)].value)
            category = str(sheet["E" + str(i)].value)
            ##  NEWS TYPE: 0 MULTIMODAL 1 IMAGE 2 TEXT
            # if not is_use_unimodal or "multi" in category:
            #     category = 0
            # elif "image" in category:
            #     category = 1
            # else:
            #     category = 2
            # imgs = images_name.split('|')
            record = {}
            record["images"] = images_name
            record["label"] = label
            record["content"] = content
            record["subfolder"] = "{}_{}".format(
                self.dataset_name, "train" if is_train else "test"
            )
            record["category"] = category
            self.label_dict.append(record)
            if label == 1 and self.is_train:
                for times in range(self.duplicate_fake_times):
                    self.label_dict.append(record)
        print(f"Skipped Num {skipped_num}")
        self.not_valid_set = set()

        # # AMBIGUITY LEARNING
        # if self.dataset_name == "gossip" and self.is_train and self.with_ambiguity:
        #     wb = openpyxl.load_workbook(
        #         self.root_path + "/{}_{}.xlsx".format(self.dataset_name, "train")
        #     )
        #     sheetnames = wb.sheetnames
        #     sheet = wb[sheetnames[0]]
        #     rows = sheet.max_row
        #     for i in tqdm(range(2, rows + 1)):
        #         images_name = str(sheet["C" + str(i)].value)
        #         label = int(sheet["D" + str(i)].value)
        #         # 1 stands for non-related
        #         content = str(sheet["B" + str(i)].value)
        #         category = str(sheet["E" + str(i)].value)
        #         ##  NEWS TYPE: 0 MULTIMODAL 1 IMAGE 2 TEXT
        #         # if not is_use_unimodal or "multi" in category:
        #         #     category = 0
        #         # elif "image" in category:
        #         #     category = 1
        #         # else:
        #         #     category = 2
        #         # imgs = images_name.split('|')
        #         record = {}
        #         record["images"] = images_name
        #         record["label"] = label
        #         record["content"] = content
        #         record["subfolder"] = "{}_{}".format(
        #             self.dataset_name, "train" if is_train else "test"
        #         )
        #         record["category"] = category
        #         self.label_ambiguity.append(record)

        assert len(self.label_dict) != 0, "Error: GT path is empty."

    def __getitem__(self, index):

        GT_size = self.image_size  # 这个自己指定一下
        find_path = False
        # get GT image
        while not find_path:
            record = self.label_dict[index]
            images, label, content = (
                record["images"],
                record["label"],
                record["content"],
            )

            GT_path = images  # imgs[np.random.randint(0,len(imgs))]
            img_GT = None
            # try:
            GT_path = "{}/{}/{}/{}".format(
                self.root_path, "Images", record["subfolder"], GT_path
            )
            if not GT_path in self.not_valid_set:

                img_GT = cv2.imread(GT_path, cv2.IMREAD_COLOR)  # util.read_img(GT_path)
                if img_GT is None:
                    img_GT = Image.open(GT_path)
                    img_GT = self.resize_and_to_tensor(img_GT).float()
                    if img_GT.shape[0] == 1:
                        img_GT = img_GT.expand(3, -1, -1)
                    elif img_GT.shape[0] == 4:
                        img_GT = img_GT[:3, :, :]
                else:
                    img_GT = img_GT.astype(np.float32) / 255.0
                    if img_GT.ndim == 2:
                        img_GT = np.expand_dims(img_GT, axis=2)
                    # some images have 4 channels
                    if img_GT.shape[2] > 3:
                        img_GT = img_GT[:, :, :3]

                    img_GT = util.channel_convert(img_GT.shape[2], "RGB", [img_GT])[0]
                    H_origin, W_origin, _ = img_GT.shape

                    ###### directly resize instead of crop
                    img_GT = cv2.resize(
                        np.copy(img_GT),
                        (GT_size, GT_size),
                        interpolation=cv2.INTER_LINEAR,
                    )

                    orig_height, orig_width, _ = img_GT.shape
                    H, W, _ = img_GT.shape

                    # BGR to RGB, HWC to CHW, numpy to tensor
                    if img_GT.shape[2] == 3:
                        img_GT = img_GT[:, :, [2, 1, 0]]

                    img_GT = torch.from_numpy(
                        np.ascontiguousarray(np.transpose(img_GT, (2, 0, 1)))
                    ).float()

                    if (
                        H_origin < 100
                        or W_origin < 100
                        or H_origin / W_origin < 0.33
                        or H_origin / W_origin > 3
                    ):  #'text' in category:
                        # print(f"Unimodal text detected {H_origin} {W_origin}. Set as zero matrix")
                        find_path = False
                        # self.not_valid_set.add(GT_path)
                        # img_GT = torch.zeros_like(img_GT)
                    elif len(content) < 10:  #'image' in category:
                        # print("Unimodal image detected. Set as \"No image provided for this news\"")
                        # content = "No image provided for this news"
                        find_path = False
                        self.not_valid_set.add(GT_path)
                    else:
                        find_path = True

            index = np.random.randint(0, len(self.label_dict))

            # except Exception:
            #     print("[Exception] load image error at {}. Using a zero-matrix instead")
            #     img_GT = torch.zeros((3, GT_size, GT_size))

        # AMBIGUITY: SAME, REPEATED CODE
        return (content, img_GT, img_GT, label, 0), (GT_path)

    def __len__(self):
        return len(self.label_dict)

    def to_tensor(self, img):
        img = Image.fromarray(img)
        img_t = F.to_tensor(img).float()
        return img_t
