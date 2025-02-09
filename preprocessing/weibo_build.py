import time
import json
from tkinter import *
from PIL import Image, ImageTk
import wget
import copy
import os
import openpyxl
import pandas as pd
import numpy as np
import paramiko
from scp import SCPClient
from tqdm import tqdm
import datetime
import random
from tkinter.filedialog import askdirectory, askopenfilename
import shutil
import cv2
import pickle
import random
random.seed(1000)

from variables import rumor_root, weibo_fake_root, weibo_real_root, weibo21_fake_root, weibo21_real_root, \
    get_workbook


def generate_ambiguity():
    records_list = []
    dataset_name = 'weibo'
    granted_positive_times, granted_negative_times = 1, 1
    xlsxs = ["../dataset/weibo/origin_do_not_modify/train_datasets_WWW_new.xlsx",
             "../dataset/weibo/origin_do_not_modify/test_datasets_origin.xlsx"]
    for xlsx in xlsxs:
        num_real_count, num_fake_count = 0, 0
        sheet_rumor, rows_rumor = get_workbook(xlsx)
        valid_list_positive = {} #set([i for i in range(2,rows_rumor+1)])
        for i in range(2,rows_rumor+1):
            valid_list_positive[i] = granted_positive_times
        valid_list_negative_text = {} #set([i for i in range(2, rows_rumor + 1)])
        for i in range(2,rows_rumor+1):
            valid_list_negative_text[i] = granted_negative_times
        if dataset_name=='gossip':
            THRESH = 5000
            LIST = ['C','D','B','E']
        else:
            THRESH = 3500 if "train" in xlsx else 500
            LIST = ['F', 'C', 'E', 'X']
        # GET 10000 PAIRS

        while (num_real_count<THRESH or num_fake_count<THRESH) and len(valid_list_negative_text)>0 and len(valid_list_positive)>0:  # title label content image B C E F
            # NOTE: 1 REPRESENTS REAL NEWS
            ########### Generate positive sample ####################################
            records = {}
            iP = random.choice(list(valid_list_positive))

            images_name = str(sheet_rumor[LIST[0] + str(iP)].value)
            label = int(sheet_rumor[LIST[1] + str(iP)].value)
            content = str(sheet_rumor[LIST[2] + str(iP)].value)
            if label==0:
                del valid_list_positive[iP]
                continue

            full_valid_name = []
            new_images = images_name.split('|')
            for new_image in new_images:
                is_found_image = cv2.imread("/home/yifan40/multimodalFNDData/data/weibo/" + new_image)
                if is_found_image is not None:
                    image_width, image_height = is_found_image.shape[0], is_found_image.shape[1]
                    ## CONDUCT FILTERING OUT INVALID NEWS IF NOT NO_FILTER_OUT
                    ratio = float(image_width) / image_height
                    if ratio>3 or ratio<0.33 or image_width < 100 or image_height < 100:
                        pass
                        # news_type = "text"
                        # print("Size {} too small {} skip..".format(is_found_image.shape,new_image))
                        # num_too_small += 1
                    else:
                        full_valid_name.append(new_image)
            # collect the news as positive if both image and text satisfies
            images_name = ''
            if len(full_valid_name)!=0:
                images_name = '|'.join(full_valid_name)
            if len(images_name)!=0 and len(content)>=10:
                images_name = '|'.join(full_valid_name)
                records['content'] = content
                records['image'] = images_name
                records['label'] = 0  # using soft_label
                records['debug'] = "{} {}".format(iP, iP)
                records_list.append(records)
                records = {}
                num_real_count += 1
                print(f"Got {num_real_count} positive samples")
            del valid_list_positive[iP]


            if len(images_name)!=0 and len(valid_list_negative_text)>0 and num_fake_count<THRESH:
                label, iN, content = 0, -1, ''
                while iN == -1 and len(valid_list_negative_text)>0:
                    iN = random.choice(list(valid_list_negative_text))
                    label = int(sheet_rumor[LIST[1] + str(iN)].value)
                    content = str(sheet_rumor[LIST[2] + str(iN)].value)
                    if label==0 or iN==iP or len(content)<10:
                        del valid_list_negative_text[iN]
                        iN = -1

                if iN!=-1:
                    # collect the news as negative if both image and text satisfies
                    records['content'] = content
                    records['image'] = images_name
                    records['label'] = 1
                    records['debug'] = "{} {}".format(iN,iP)
                    records_list.append(records)
                    records = {}
                    num_fake_count += 1
                    print(f"Got {num_fake_count} negative samples")
                    del valid_list_negative_text[iN]


        print("{} {} {}".format(num_real_count,num_fake_count,len(records_list)))
    ambiguity_excel = "../dataset/{}/origin_do_not_modify/{}_{}.xlsx".format(dataset_name, dataset_name, "train_ambiguity_new")
    df = pd.DataFrame(records_list)
    df.to_excel(ambiguity_excel)

def weibo_text_to_excel(is_train="both",conduct_download=True):
    is_train_list = ["train","test"]
    num_too_small = 0
    # weibo_train_id, weibo_val_id, weibo_test_id = None ,None ,None
    num_length, num_cv_error, num_no_valid = 0, 0, 0
    train_list, val_list, test_list = [], [], []
    for is_train in is_train_list:
        non_rumor_text = '/home/yifan40/multimodalFNDData/data/weibo/tweets/{}_nonrumor.txt'.format(is_train)
        rumor_text = '/home/yifan40/multimodalFNDData/data/weibo/tweets/{}_rumor.txt'.format(is_train)
        text_lists = [rumor_text, non_rumor_text]
        results = []
        for i in tqdm(range(len(text_lists))):
            text_list = text_lists[i]
            f = open(text_list,encoding='UTF-8')  # 返回一个文件对象
            line = f.readline()  # 调用文件的 readline()方法
            line_num,results_dict = 0, {}
            while line:  # 如果每行是json的字符串形式  json.loads(line) # line_json["filename"])
                if line_num%3==0: # 无用信息
                    results_dict['source'] = line
                elif line_num%3==1: # img
                    # imgs = line.split('|')
                    # imgs = [img[img.rfind('/')+1] for img in imgs]
                    results_dict['images'] = line #imgs
                else: # content
                    results_dict['content'] = line
                    results_dict['label'] = i
                    results.append(results_dict)
                    results_dict = {}
                line = f.readline()
                line_num+=1
            f.close()

        df = pd.DataFrame(results)
        filename = '/home/yifan40/multimodalFNDData/data/weibo/origin_do_not_modify/{}_datasets_origin.xlsx'.format(is_train)
        df.to_excel(filename)
        print(f"File saved at {filename}")

        folder1, folder2 = "/home/yifan40/multimodalFNDData/data/weibo/nonrumor_images", "/home/yifan40/multimodalFNDData/data/weibo/rumor_images"
        images1, images2 = set(os.listdir(folder1)), set(os.listdir(folder2))
        images_set = images1.union(images2)
        news_xlsxs = []
        if is_train=="train" or is_train=='both':
            news_xlsxs.append('../dataset/weibo/origin_do_not_modify/train_datasets_origin.xlsx')
        if is_train == "test" or is_train == 'both':
            news_xlsxs.append('../dataset/weibo/origin_do_not_modify/test_datasets_origin.xlsx')

        for news_xlsx in news_xlsxs:
            wb = openpyxl.load_workbook(news_xlsx)
            sheetnames = wb.sheetnames
            sheet = wb[sheetnames[0]]
            rows = sheet.max_row
            for i in tqdm(range(2, rows + 1)):
                source = sheet['B' + str(i)].value
                title, full_valid_name = '', []
                content = sheet['D' + str(i)].value
                if content is None or len(content) <=10:
                    print("Found length not satisfy {}".format(content))
                    num_length += 1
                    continue
                image_names = sheet['C' + str(i)].value
                label = int(sheet['E' + str(i)].value)
                if image_names is not None and len(image_names) != 0:
                    image_name_list = image_names.split('|')[:-1]
                    for image_name in image_name_list:
                        full_image_name = image_name.lstrip(' ').rstrip(' ')  # http url
                        short_name = full_image_name[full_image_name.rfind('/') + 1:]
                        ## if not found, we tempt to download... if cannot download, we skip...
                        if short_name not in images_set:
                            if conduct_download:
                                try:
                                    wget.download(full_image_name, "./temp_images/"+("rumor_images/" if label == 0 else "nonrumor_images/") + short_name)
                                    # image_name = image_name + "rumor_images/" + short_name + "|"
                                    print("Download ok. {}".format(full_image_name))
                                    images_set.add(short_name)
                                except Exception:
                                    print("Download error. {}".format(full_image_name))
                            else:
                                print("Do not download. {}".format(full_image_name))

                        if short_name in images_set:
                            new_image = ("rumor_images/" if label == 0 else "nonrumor_images/") + short_name
                            is_found_image = cv2.imread("/home/yifan40/multimodalFNDData/data/weibo/"+new_image)
                            if is_found_image is not None:
                                image_width, image_height = is_found_image.shape[0], is_found_image.shape[1]
                                ## CONDUCT FILTERING OUT INVALID NEWS IF NOT NO_FILTER_OUT
                                ratio = float(image_width) / image_height
                                # if ratio>3 or ratio<0.33 or image_width < 100 or image_height < 100:
                                #     news_type = "text"
                                #     print("Size {} too small {} skip..".format(is_found_image.shape,new_image))
                                #     num_too_small += 1
                                # else:
                                full_valid_name.append(new_image)
                if len(full_valid_name)!=0:
                    full_valid_name = '|'.join(full_valid_name)
                    record = {}
                    record['title'], record['label'], record['source'], record['content'], record['image'] \
                        = title, label, source, content, full_valid_name
                    ## send ''category'' to the data loader, so that it can replace some of the useless modalities into placeholders.
                    # record['category'] =
                    query_key = source.split('|')[0]
                    if "train" in news_xlsx:
                        train_list.append(record)
                    else:
                        test_list.append(record)
                else:
                    print("No valid image found {}".format(content))
                    num_no_valid += 1

    # if "train" in is_train:
    filename = '../dataset/weibo/origin_do_not_modify/train_datasets_WWW_new.xlsx'
    df = pd.DataFrame(train_list)
    df.to_excel(filename)
    print(f"File saved at {filename}")
    filename = '../dataset/weibo/origin_do_not_modify/test_datasets_WWW_new.xlsx'
    df = pd.DataFrame(test_list)
    df.to_excel(filename)
    print(f"File saved at {filename}")
    print("length {} CV_Error {} No_valid {}".format(num_length,num_cv_error, num_no_valid))
    print("Train_set {} Test_set {} Val_set {}".format(len(train_list),len(test_list),len(val_list)))
    print(f"Size not valid {num_too_small}")

if __name__ == '__main__':
    weibo_text_to_excel('both')